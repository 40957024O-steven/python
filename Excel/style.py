import openpyxl
from openpyxl.styles import Font
from openpyxl import Workbook, load_workbook


wb = Workbook()
ws = wb.active

ws.column_dimensions['A'].width = 20.0 #調整A欄寬 = 20
ws.row_dimensions[1].height = 40 # 調整行1高
ws["A1"].font = Font (bold = True , color= "000000FF") #將A1格子字體設為粗體、藍色