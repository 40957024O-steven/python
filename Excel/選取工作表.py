from openpyxl import Workbook, load_workbook #Workbook 的 W 要大寫!!

wb = load_workbook('excel.xlsx') #讀取excel
ws = wb["工作表1"]

print (ws["B2"].value) #查看某格的資料

ws["B2"].value = 123


wb.create_sheet("gm")

print(wb.sheetnames)#回傳工作表名子

wb.save("excel.xlsx") #要先關閉檔案才能存檔(只要沒存就不會去改excel)