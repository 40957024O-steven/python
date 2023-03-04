from openpyxl import Workbook, load_workbook

wb = Workbook() #創建excel
ws = wb.active
ws.title = "123" #創第一個sheet
wb.create_sheet("456") #另外再創一個sheet

ws = wb["456"]

ws.append([123,456,789,0])
ws.append([1,2,3,4,5,6])
ws.append([9,8,7,6,5,4,321])

wb.save("new.xlsx")