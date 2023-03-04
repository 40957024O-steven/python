
from openpyxl import Workbook,load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook ("new.xlsx")
ws = wb["456"]

for row in range(1,4): #讀橫排 (1~3排)
    for col in range(1,8): #讀直排 (1~7排)
        cher = get_column_letter (col)
        ws[cher+str(row)].value = cher+str(row)

wb.save("new2.xlsx") #跟原本檔案名稱不一樣的話會另存新檔(原本檔案不會變)