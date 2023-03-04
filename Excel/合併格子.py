from openpyxl import Workbook, load_workbook

wb = Workbook()
# wb = load_workbook('new2.xlsx')
ws = wb.active

ws.merge_cells("A1:E2") #合併格子
ws.unmerge_cells("A1:E2") #取消合併 (只能跟原本的範圍一樣)
wb.save("new2.xlsx")