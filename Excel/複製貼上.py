from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
wb = load_workbook ('mbs.xlsx')
ws = wb['sheet1'] #被被複製資料的Sheet
sheet = wb.active
a = sheet.max_row #複製範圍(列)
b=[]
i=0 #起始位置
while i < a:
    i = i+1
    b.insert(i, ws.cell (i,2).value) #複製的行數(i,2) ==>第二行

wbtest = load_workbook ('mbs2.0.xlsx') #貼上的位置(可以不同的excel)
ws = wbtest['sheet2'] #貼上的sheetlocation
sheet = wb.active
i = 0
while i < a:
    ws.cell (i+1,1).value = b[i] #貼上的行數(i+1,1) ==>第一行
    i = i+1
wbtest.save ('mbs2.0.xlsx')