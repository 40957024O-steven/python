import glob
from openpyxl import load_workbook,Workbook




wbnew = Workbook()
wsvl = wbnew.active
wsvl.title = '風速lidar自算'
wsvt = wbnew.create_sheet('風速探空')
wsvm = wbnew.create_sheet('風速lidar電腦算')
wsdl = wbnew.create_sheet('風向lidar自算')
wsdt = wbnew.create_sheet('風向探空')
wsdm = wbnew.create_sheet('風向lidar電腦算')
lc = 2

files = glob.glob("C:/Users/steve/Desktop/python相關資料/觀測rowdata/統計/*")
for file in files:  
    day = file[47:60]


# day = '2019-01-01_00'

    wbold = load_workbook("C:/Users/steve/Desktop/python相關資料/觀測rowdata/統計/"+day+'.xlsx')
    wsv = wbold['風速']
    wsd = wbold['風向']
    hivmax = wsv.max_row
    hidmax = wsd.max_row
    # print(himax)
    for i in range(2,hivmax+1):
        wsvl.cell(i,1).value = wsv.cell(i,1).value
        wsvl.cell(i,lc).value = wsv.cell(i,2).value

        wsvt.cell(i,1).value = wsv.cell(i,1).value
        wsvt.cell(i,lc).value = wsv.cell(i,3).value

        wsvm.cell(i,1).value = wsv.cell(i,1).value
        wsvm.cell(i,lc).value = wsv.cell(i,6).value        
    for i in range(2,hidmax+1):
        wsdl.cell(i,1).value = wsd.cell(i,1).value
        wsdl.cell(i,lc).value = wsd.cell(i,2).value

        wsdt.cell(i,1).value = wsd.cell(i,1).value
        wsdt.cell(i,lc).value = wsd.cell(i,3).value
    
        wsdm.cell(i,1).value = wsd.cell(i,1).value
        wsdm.cell(i,lc).value = wsd.cell(i,6).value 

    wsvl.cell(1,lc).value = day
    wsvt.cell(1,lc).value = day
    wsvm.cell(1,lc).value = day    

    wsdl.cell(1,lc).value = day
    wsdt.cell(1,lc).value = day
    wsdm.cell(1,lc).value = day

    lc +=1
wbnew.save("C:/Users/steve/Desktop/python相關資料/觀測rowdata/統計.xlsx")