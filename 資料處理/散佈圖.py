from subprocess import call
import matplotlib.pylab as plt
from openpyxl import workbook,load_workbook
import matplotlib.ticker as ticker
wb = load_workbook('randomnb.xlsx')
ws = wb['工作表1']

X = []
Y = []

for i in range(1,100):
    X.append(ws.cell(i,1).value)
    Y.append(ws.cell(i,2).value)

fig = plt.figure() #取得一張空白的map
 
axes = fig.add_subplot(1,1,1) #設定chunk數 add_subplot(行,列,第幾個chunk)
# axes.set_xlim(1,10.1) #簡單的X軸設定
axes.set_xticks(range(0,105,5)) #較為詳細的X軸設定  range(上限,下限,單位變化))(上限要+1)
axes.set_yticks(range(0,11,1))
axes.xaxis.set_major_locator(ticker.LinearLocator(27)) #刻度set

axes.scatter(X,Y,color = 'blue',s=10,label='123') #散部圖 (s=點大小,alpha=透明度,marker = 點樣式,label = 圖例)
plt.show() #關掉後才會繼續執行程式

