import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.writer.excel import ExcelWriter
import numpy as np

wb = load_workbook('123.xlsx')
ws = wb['random']


Z = []
Z0 = []

def split_list(l, n):

  for idx in range(0, len(l), n):
    yield l[idx:idx+n]

for i in range(1,10001):

    Z0.insert(i,ws.cell(i,1).value)
X = range(100)
Y = range(100)
Z = list(split_list(Z0,100))

# print(Z)
plt.contour(X,Y,Z,levels = 100)
plt.colorbar()
plt.show()
